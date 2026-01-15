from django.shortcuts import render, redirect, get_object_or_404

from django.db.models import Q
from user.models import UserInfo
from user.forms import UserInfoForm
from django.contrib.auth.hashers import make_password  # 密码加密

# 以下是创建角色组和初始用户的代码
from django.contrib.auth.models import User, Group, Permission

from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.shortcuts import redirect
from user.models import UserInfo
from booking.models import Booking, ApprovalRecord
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from labadmin.views import handle_approval

# Create your views here.
# ---------------------- 负责人视图 ----------------------
@login_required
def manager_home(request):
    """负责人首页 - 仅允许负责人访问"""
    # 验证用户身份：确保request.user是正确的用户
    if not request.user.is_authenticated:
        from django.contrib.auth import logout
        logout(request)
        messages.error(request, '登录已过期，请重新登录！')
        return redirect('user_login')
    
    # 验证用户是否是负责人（严格检查，不允许管理员访问）
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    
    # 关键修复：如果是管理员但不是负责人，必须重定向到管理员首页
    if is_admin and not is_manager and not request.user.is_superuser:
        messages.info(request, '您已切换到管理员界面')
        return redirect('admin_home')
    
    # 如果不是负责人，重定向到对应首页或登录页
    if not is_manager:
        # 如果是普通用户，重定向到普通用户首页
        try:
            from user.models import UserInfo
            user_info = UserInfo.objects.get(auth_user=request.user)
            return redirect('user_home')
        except:
            pass
        # 否则重定向到登录页
        messages.error(request, '您不是实验室负责人，无权访问此页面！')
        return redirect('user_login')
    
    return render(request, 'manager/home.html')

@login_required
@csrf_protect
def booking_approve(request):
    """设备预约审批（负责人） - 仅允许负责人访问"""
    # 校验是否是负责人（严格检查，不允许管理员访问负责人审批页面）
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 关键修复：如果是管理员但不是负责人，重定向到管理员审批页面
    if is_admin and not is_manager and not request.user.is_superuser:
        messages.info(request, '管理员请使用管理员审批页面')
        return redirect('booking_approve')  # 管理员审批页面
    
    # 如果不是负责人，重定向到对应首页或登录页
    if not is_manager:
        # 如果是普通用户，重定向到普通用户首页
        try:
            from user.models import UserInfo
            user_info = UserInfo.objects.get(auth_user=request.user)
            return redirect('user_home')
        except:
            pass
        messages.error(request, '你无审批权限！')
        return redirect('user_login')
    
    # 获取筛选条件
    user_type_filter = request.GET.get('user_type', 'all')
    
    # ========== 核心修改：精准限定各角色的查询范围 ==========
    if is_admin:
        # 管理员：只看「待管理员审批（pending）」的申请（已批准的不再显示）
        bookings = Booking.objects.filter(
            status='pending'  # 仅待管理员审批
        ).order_by('-create_time')
    else:  # 实验室负责人
        # 负责人：只看「管理员已批准（admin_approved）」的申请（且仅校外人员）
        bookings = Booking.objects.filter(
            status='admin_approved',  # 仅管理员已批准的
            applicant__user_type='external'  # 仅校外人员（符合原有业务规则）
        ).order_by('-create_time')
    
    # 筛选用户类型（仅对管理员生效，负责人默认只看校外人员，筛选不影响）
    if user_type_filter != 'all':
        if is_admin:  # 管理员可筛选所有类型
            bookings = bookings.filter(applicant__user_type=user_type_filter)
        # 负责人：强制限定为校外人员，忽略其他筛选（避免误看非校外人员申请）
        elif user_type_filter == 'external':
            bookings = bookings.filter(applicant__user_type='external')
    
    # 处理审批操作（原有逻辑不变）
    if request.method == 'POST':
        # 单个审批
        if 'approve' in request.POST:
            booking_id = request.POST.get('approve')
            handle_approval(request, booking_id, 'approve')
        elif 'reject' in request.POST:
            booking_id = request.POST.get('reject')
            handle_approval(request, booking_id, 'reject')
        
        # 批量审批（简化版）
        elif 'batch_approve' in request.POST or 'batch_reject' in request.POST:
            booking_ids = request.POST.getlist('booking_ids')
            action = 'approve' if 'batch_approve' in request.POST else 'reject'
            for booking_id in booking_ids:
                handle_approval(request, booking_id, action)
        
        # 保留查询参数（筛选条件），避免跳转错误
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        redirect_url = reverse('manager_booking_approve')
        query_params = request.GET.copy()
        if query_params:
            redirect_url += '?' + query_params.urlencode()
        return HttpResponseRedirect(redirect_url)
    
    context = {
        'bookings': bookings,
        'user_type_filter': user_type_filter,
        'is_admin': is_admin,
        'is_manager': is_manager
    }
    return render(request, 'manager/booking_approve.html', context)

@login_required
@csrf_protect
def manager_report_stat(request):
    """负责人报表统计页面 - 仅允许负责人访问"""
    # 验证用户是否是负责人（严格检查，不允许管理员访问）
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 关键修复：如果是管理员但不是负责人，重定向到管理员报表页面
    if is_admin and not is_manager and not request.user.is_superuser:
        messages.info(request, '管理员请使用管理员报表页面')
        return redirect('report_stat')
    
    # 如果不是负责人，重定向到对应首页或登录页
    if not is_manager:
        # 如果是普通用户，重定向到普通用户首页
        try:
            from user.models import UserInfo
            user_info = UserInfo.objects.get(auth_user=request.user)
            return redirect('user_home')
        except:
            pass
        messages.error(request, '您不是实验室负责人，无权访问此页面！')
        return redirect('user_login')
    
    from labadmin.models import Report
    from labadmin.views import generate_report_data
    from datetime import datetime, timedelta, date
    from decimal import Decimal

    # 获取已生成的报表列表
    reports = Report.objects.all().order_by('-generated_at')[:20]
    
    # 获取筛选条件
    report_type_filter = request.GET.get('report_type', '')
    date_filter = request.GET.get('date', '')
    
    if report_type_filter:
        reports = reports.filter(report_type=report_type_filter)
    
    # 处理报表生成请求
    if request.method == 'POST' and 'generate' in request.POST:
        report_type = request.POST.get('report_type')
        date_input = request.POST.get('date_input', '').strip()
        start_date_input = request.POST.get('start_date', '').strip()
        end_date_input = request.POST.get('end_date', '').strip()
        
        # 自定义时间段报表需要起始日期和结束日期
        if report_type == 'custom':
            if not start_date_input or not end_date_input:
                messages.error(request, '自定义时间段报表需要填写起始日期和结束日期！')
                return redirect('manager_report_stat')
        elif not date_input:
            messages.error(request, '请选择报表类型和日期！')
            return redirect('manager_report_stat')
        
        try:
            # 解析日期
            if report_type == 'week':
                # 周报表：输入日期所在周的周一和周日
                input_date = datetime.strptime(date_input, '%Y-%m-%d').date()
                start_date = input_date - timedelta(days=input_date.weekday())
                end_date = start_date + timedelta(days=6)
                report_name = f"{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')} 周报表"
            elif report_type == 'month':
                # 月报表：输入日期所在月的第一天和最后一天
                # 处理 YYYY-MM 格式
                if len(date_input) == 7 and date_input.count('-') == 1:
                    year, month = map(int, date_input.split('-'))
                else:
                    # 尝试解析为日期
                    input_date = datetime.strptime(date_input, '%Y-%m-%d').date()
                    year, month = input_date.year, input_date.month
                start_date = date(year, month, 1)
                if month == 12:
                    end_date = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(year, month + 1, 1) - timedelta(days=1)
                report_name = f"{year}年{month:02d}月报表"
            elif report_type == 'year':
                # 年报表：输入日期所在年的1月1日和12月31日
                year = int(date_input)
                start_date = date(year, 1, 1)
                end_date = date(year, 12, 31)
                report_name = f"{year}年报表"
            elif report_type == 'custom':
                # 自定义时间段报表：使用用户指定的起始日期和结束日期
                start_date = datetime.strptime(start_date_input, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_input, '%Y-%m-%d').date()
                
                # 验证日期范围
                if start_date > end_date:
                    messages.error(request, '起始日期不能晚于结束日期！')
                    return redirect('manager_report_stat')
                
                report_name = f"{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')} 自定义报表"
            else:
                messages.error(request, '无效的报表类型！')
                return redirect('manager_report_stat')
            
            # 检查是否已存在相同报表（只检查相同类型的，自定义报表可以重复生成）
            if report_type != 'custom':
                existing_report = Report.objects.filter(
                    report_type=report_type,
                    start_date=start_date,
                    end_date=end_date
                ).first()
                
                if existing_report:
                    if existing_report.generated_by:
                        messages.info(request, f'该时间段报表已存在（手动生成），已为您加载：{existing_report.report_name}')
                    else:
                        messages.info(request, f'该时间段报表已存在（系统自动生成），已为您加载：{existing_report.report_name}')
                    return redirect(f'manager_report_stat?view={existing_report.id}')
            
            # 生成报表数据
            report_data = generate_report_data(report_type, start_date, end_date)
            
            # 创建报表记录
            report = Report.objects.create(
                report_type=report_type,
                report_name=report_name,
                start_date=start_date,
                end_date=end_date,
                report_data=report_data,
                total_bookings=report_data['summary']['total_bookings'],
                total_devices=report_data['summary']['total_devices'],
                total_users=report_data['summary']['total_users'],
                total_revenue=Decimal(str(report_data['summary']['total_revenue'])),
                generated_by=request.user
            )
            
            messages.success(request, f'报表生成成功：{report_name}')
            return redirect(f'manager_report_stat?view={report.id}')
            
        except ValueError as e:
            messages.error(request, f'日期格式错误：请检查日期格式是否正确！')
            return redirect('manager_report_stat')
        except Exception as e:
            messages.error(request, f'生成报表失败：{str(e)}')
            return redirect('manager_report_stat')
    
    # 查看报表详情
    report_id = request.GET.get('view')
    current_report = None
    if report_id:
        try:
            current_report = Report.objects.get(id=report_id)
        except Report.DoesNotExist:
            messages.error(request, '报表不存在！')
    
    context = {
        'reports': reports,
        'current_report': current_report,
        'report_type_filter': report_type_filter,
        'date_filter': date_filter,
    }
    return render(request, 'manager/report_stat.html', context)

@login_required
def manager_delete_report(request, report_id):
    """删除报表"""
    from labadmin.models import Report
    
    try:
        report = Report.objects.get(id=report_id)
        report_name = report.report_name
        report.delete()
        messages.success(request, f'报表【{report_name}】已成功删除！')
    except Report.DoesNotExist:
        messages.error(request, '报表不存在！')
    except Exception as e:
        messages.error(request, f'删除报表失败：{str(e)}')
    
    # 保留查询参数，避免跳转错误
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    redirect_url = reverse('manager_report_stat')
    # 保留原有的筛选条件
    query_params = request.GET.copy()
    if 'view' in query_params:
        del query_params['view']  # 删除view参数，因为报表已删除
    if query_params:
        redirect_url += '?' + query_params.urlencode()
    return HttpResponseRedirect(redirect_url)

@login_required
def manager_export_report_csv(request, report_id):
    """负责人导出报表为Excel文件（.xlsx）"""
    from django.http import HttpResponse
    from labadmin.models import Report
    import re
    
    report = get_object_or_404(Report, id=report_id)
    data = report.get_report_data()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "报表"
    
    # 写入报表基本信息
    # 转换datetime为naive datetime（移除时区信息）
    generated_at = report.generated_at.replace(tzinfo=None) if report.generated_at else None
    
    ws.append(['报表名称', report.report_name])
    ws.append(['报表类型', report.get_report_type_display()])
    ws.append(['统计时间', f'{report.start_date.strftime("%Y-%m-%d")} 至 {report.end_date.strftime("%Y-%m-%d")}'])
    ws.append(['生成时间', generated_at])
    ws.append(['生成人', report.generated_by.username if report.generated_by else '系统自动'])
    ws.append([])  # 空行
    
    # 设置生成时间为日期格式
    if generated_at:
        ws.cell(row=4, column=2).number_format = 'yyyy-mm-dd hh:mm:ss'
    
    # 写入汇总统计
    ws.append(['汇总统计'])
    ws.append(['总预约次数', data['summary']['total_bookings']])
    ws.append(['已审批通过', data['summary']['approved_count']])
    ws.append(['总收入（元）', data['summary']['total_revenue']])
    ws.append(['设备总数', data['summary']['total_devices']])
    ws.append(['用户总数', data['summary']['total_users']])
    ws.append([])  # 空行
    
    # 写入设备使用统计
    ws.append(['设备使用统计'])
    headers = ['设备编号', '设备型号', '预约次数', '使用时长（小时）', '使用率（%）', '校外收费（元）']
    ws.append(headers)
    
    # 设置表头样式
    header_row = ws.max_row
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for device in data.get('device_usage', []):
        ws.append([
            device['device_code'],
            device['device_model'],
            device['booking_count'],
            device['usage_hours'],
            f"{device['usage_rate']}%",
            device['revenue']
        ])
    ws.append([])  # 空行
    
    # 写入用户类型统计
    ws.append(['用户类型统计'])
    headers = ['用户类型', '预约次数', '用户数量']
    ws.append(headers)
    
    # 设置表头样式
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for stat in data.get('user_type_stats', []):
        user_type = stat.get('applicant__user_type', '')
        user_type_display = {
            'student': '校内学生',
            'teacher': '校内教师',
            'external': '校外人员'
        }.get(user_type, user_type)
        ws.append([
            user_type_display,
            stat['booking_count'],
            stat['user_count']
        ])
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 创建响应
    # 清理文件名中的特殊字符
    import re
    safe_filename = re.sub(r'[<>:"/\\|?*]', '_', report.report_name)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{report.id}_{safe_filename}.xlsx"'
    
    wb.save(response)
    return response

# -----------------------s--- 1. 用户列表（含搜索、筛选） --------------------------
@login_required
@csrf_protect
def user_manage(request):
    """
    用户管理主页面：展示所有用户，支持按姓名/编号搜索、按类型筛选
    对应路径：/manager/user/manage/
    """
    # 1. 处理筛选和搜索（原有逻辑不变）
    user_type = request.GET.get('user_type', '')
    keyword = request.GET.get('keyword', '')
    approval_status = request.GET.get('approval_status', '')
    # 使用annotate添加借用次数统计
    from django.db.models import Count
    user = UserInfo.objects.annotate(booking_count=Count('booking'))
    # 获取所有管理员用户（不关联UserInfo的User对象）
    admin_users = User.objects.filter(groups__name='设备管理员').exclude(
        id__in=UserInfo.objects.values_list('auth_user_id', flat=True)
    )
    
    if user_type and user_type in ['student', 'teacher', 'external']:
        user = user.filter(user_type=user_type)
    elif user_type == 'admin':
        # 筛选管理员用户，不显示UserInfo用户
        user = UserInfo.objects.none()  # 空查询集
    if keyword:
        if user_type != 'admin':
            user = user.filter(Q(name__icontains=keyword) | Q(user_code__icontains=keyword))
        else:
            admin_users = admin_users.filter(Q(username__icontains=keyword) | Q(first_name__icontains=keyword))
    if approval_status == 'pending':
        user = user.filter(approval_status='pending')
    
    # 2. 处理新增用户（POST请求）【核心修改：用户名=用户编号，密码=用户编号】
    # 注意：先检查是否是审核操作，如果不是，再处理新增用户
    if request.method == 'POST' and 'approve_user' not in request.POST:
        # 检查是否是创建管理员
        create_admin = request.POST.get('create_admin', '') == 'true'
        
        if create_admin:
            # 创建管理员用户（不创建UserInfo）
            admin_username = request.POST.get('admin_username', '').strip()
            admin_name = request.POST.get('admin_name', '').strip()
            
            if not admin_username or not admin_name:
                messages.error(request, '管理员用户名和姓名不能为空！')
            elif User.objects.filter(username=admin_username).exists():
                messages.error(request, f'用户名【{admin_username}】已存在！')
            else:
                # 创建管理员用户
                admin_user = User.objects.create(
                    username=admin_username,
                    password=make_password(admin_username),  # 初始密码=用户名
                    first_name=admin_name,
                    is_active=True,
                    is_staff=True  # 管理员需要is_staff=True
                )
                # 添加到设备管理员组
                admin_group = Group.objects.get(name='设备管理员')
                admin_user.groups.add(admin_group)
                admin_user.save()
                
                messages.success(request, f'管理员【{admin_name}】创建成功！用户名：{admin_username}，初始密码：{admin_username}')
                # 保留查询参数
                from django.http import HttpResponseRedirect
                from django.urls import reverse
                redirect_url = reverse('user_manage')
                query_params = request.GET.copy()
                if query_params:
                    redirect_url += '?' + query_params.urlencode()
                return HttpResponseRedirect(redirect_url)
        else:
            # 创建普通用户（原有逻辑）
            form = UserInfoForm(request.POST)
            if form.is_valid():
                # 先保存UserInfo（不提交到数据库）
                user_info = form.save(commit=False)
                
                # 【关键修改】用户名 = 用户编号（学号/工号），密码 = 用户编号
                username = user_info.user_code  # 账号用编号（如20230001/T001/O001）
                password = make_password(user_info.user_code)  # 密码用编号（加密）
                
                # 检查用户名是否已存在（用户编号本身已设置unique=True，此处双重保险）
                if User.objects.filter(username=username).exists():
                    # 理论上不会触发，因为user_code是唯一的
                    form.add_error('user_code', '该用户编号已作为登录账号存在！')
                    return render(request, 'manager/user_manage.html', {
                        'users': user,
                        'admin_users': admin_users if user_type == 'admin' else [],
                        'keyword': keyword,
                        'user_type': user_type,
                        'form': form,
                        'total_users': user.count() + (admin_users.count() if user_type == 'admin' else 0),
                        'active_users': user.filter(is_active=True).count() + (admin_users.filter(is_active=True).count() if user_type == 'admin' else 0),
                        'inactive_users': user.filter(is_active=False).count() + (admin_users.filter(is_active=False).count() if user_type == 'admin' else 0),
                        'users_with_bookings': user.filter(booking_count__gt=0).count(),
                        'pending_approvals': UserInfo.objects.filter(approval_status='pending').count(),
                    })
                
                # 创建登录账号
                auth_user = User.objects.create(
                    username=username,          # 账号=用户编号
                    password=password,          # 密码=用户编号（加密）
                    first_name=user_info.name,  # 姓名（可选）
                    is_active=user_info.is_active  # 借用资格=账号是否激活
                )
                
                # 关联登录账号到UserInfo，并加入普通用户组
                user_info.auth_user = auth_user
                user_info.save()
                
                # 可选：将普通用户加入「普通用户」组
                user_group = Group.objects.get(name='普通用户')
                auth_user.groups.add(user_group)
                auth_user.save()
                
                messages.success(request, f'用户【{user_info.name}】创建成功！')
                # 保留查询参数（筛选条件），避免跳转错误
                from django.http import HttpResponseRedirect
                from django.urls import reverse
                redirect_url = reverse('user_manage')
                query_params = request.GET.copy()
                if query_params:
                    redirect_url += '?' + query_params.urlencode()
                return HttpResponseRedirect(redirect_url)
    else:
        form = UserInfoForm()
    
    # 3. 处理用户审核（POST请求，来自审核按钮）
    # 注意：模板中按钮的 name="approve_user" value="approve/reject"，所以这里需要检查 approve_user 的值
    if request.method == 'POST' and 'approve_user' in request.POST:
        user_id = request.POST.get('user_id')
        action = request.POST.get('approve_user')  # 获取按钮的value，应该是 'approve' 或 'reject'
        
        try:
            user_to_approve = UserInfo.objects.get(id=user_id)
            if action == 'approve':
                user_to_approve.approval_status = 'approved'
                user_to_approve.is_active = True
                if user_to_approve.auth_user:
                    user_to_approve.auth_user.is_active = True
                    user_to_approve.auth_user.save()
                user_to_approve.save()  # 确保保存状态
                messages.success(request, f'已通过用户 {user_to_approve.name} 的注册审核')
            elif action == 'reject':
                user_to_approve.approval_status = 'rejected'
                user_to_approve.is_active = False
                if user_to_approve.auth_user:
                    user_to_approve.auth_user.is_active = False
                    user_to_approve.auth_user.save()
                user_to_approve.save()  # 确保保存状态
                messages.success(request, f'已拒绝用户 {user_to_approve.name} 的注册审核')
            else:
                messages.error(request, f'无效的操作：{action}')
        except UserInfo.DoesNotExist:
            messages.error(request, '用户不存在')
        except Exception as e:
            messages.error(request, f'审批操作失败：{str(e)}')
        
        # 保留查询参数（筛选条件），避免跳转错误
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        redirect_url = reverse('user_manage')
        query_params = request.GET.copy()
        if query_params:
            redirect_url += '?' + query_params.urlencode()
        return HttpResponseRedirect(redirect_url)
    
    # 4. 准备上下文（原有逻辑不变）
    # 获取用户角色信息
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 统计信息（包含管理员）
    if user_type == 'admin':
        total_users = admin_users.count()
        active_users = admin_users.filter(is_active=True).count()
        inactive_users = admin_users.filter(is_active=False).count()
        users_with_bookings = 0  # 管理员没有借用记录
    else:
        total_users = user.count()
        active_users = user.filter(is_active=True).count()
        inactive_users = user.filter(is_active=False).count()
        users_with_bookings = user.filter(booking_count__gt=0).count()
    
    pending_approvals = UserInfo.objects.filter(approval_status='pending').count()
    
    context = {
        'users': user,
        'admin_users': admin_users if user_type == 'admin' else [],
        'keyword': keyword,
        'user_type': user_type,
        'form': form,
        'is_admin': is_admin,
        'is_manager': is_manager,
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'users_with_bookings': users_with_bookings,
        'pending_approvals': pending_approvals,
    }
    return render(request, 'manager/user_manage.html', context)

# -------------------------- 2. 编辑用户 --------------------------
@login_required
@csrf_protect
def user_edit(request, pk):
    """编辑用户信息"""
    user_info = get_object_or_404(UserInfo, pk=pk)
    
    if request.method == 'POST':
        form = UserInfoForm(request.POST, instance=user_info)
        if form.is_valid():
            user_info = form.save(commit=False)
            
            # 【关键】用户编号不可修改，保持原有编号
            original_user_info = UserInfo.objects.get(pk=user_info.pk)
            user_info.user_code = original_user_info.user_code  # 强制保持原编号
            
            # 确保登录账号的用户名与用户编号一致
            if user_info.auth_user and user_info.auth_user.username != user_info.user_code:
                user_info.auth_user.username = user_info.user_code
                user_info.auth_user.save()
            
            # 同步更新登录账号激活状态
            if user_info.auth_user:
                user_info.auth_user.is_active = user_info.is_active
                
                # 可选：如果表单填写了「重置密码为编号」，则更新密码
                reset_to_code = request.POST.get('reset_to_code', '')
                if reset_to_code:
                    user_info.auth_user.password = make_password(user_info.user_code)
                
                user_info.auth_user.save()
            
            user_info.save()
            messages.success(request, f'用户【{user_info.name}】信息已更新！')
            # 保留查询参数，避免跳转错误
            from django.http import HttpResponseRedirect
            from django.urls import reverse
            redirect_url = reverse('user_manage')
            query_params = request.GET.copy()
            if query_params:
                redirect_url += '?' + query_params.urlencode()
            return HttpResponseRedirect(redirect_url)
        else:
            # 表单验证失败
            messages.error(request, '表单填写有误，请检查后重新提交！')
    else:
        form = UserInfoForm(instance=user_info)
    
    # 获取用户角色信息
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    context = {
        'form': form,
        'user': user_info,
        'is_admin': is_admin,
        'is_manager': is_manager,
    }
    return render(request, 'manager/user_edit.html', context)

# -------------------------- 3. 删除用户 --------------------------
@login_required
def user_delete(request, pk):
    """删除用户（物理删除，确保同时删除关联的登录账号）"""
    # 获取要删除的用户扩展信息
    user_info = get_object_or_404(UserInfo, pk=pk)
    
    # 关键步骤：先删除关联的登录账号（auth_user）
    if user_info.auth_user:  # 先判断是否有关联的登录账号
        user_info.auth_user.delete()  # 主动删除Django内置User记录
    
    # 再删除自定义的UserInfo记录
    user_info.delete()
    
    # 保留查询参数，避免跳转错误
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    redirect_url = reverse('user_manage')
    query_params = request.GET.copy()
    if query_params:
        redirect_url += '?' + query_params.urlencode()
    return HttpResponseRedirect(redirect_url)

# -------------------------- 4. 快速切换用户状态（禁用/启用） --------------------------
@login_required
def user_toggle_status(request, pk):
    """快速切换用户的借用资格（无需进入编辑页）"""
    user = get_object_or_404(UserInfo, pk=pk)
    user.is_active = not user.is_active  # 取反：正常→禁用，禁用→正常
    user.save()
    
    # 同步更新登录账号的激活状态
    if user.auth_user:
        user.auth_user.is_active = user.is_active
        user.auth_user.save()
    
    messages.success(request, f'用户【{user.name}】状态已更新为{"正常" if user.is_active else "禁用"}')
    # 保留查询参数，避免跳转错误
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    redirect_url = reverse('user_manage')
    query_params = request.GET.copy()
    if query_params:
        redirect_url += '?' + query_params.urlencode()
    return HttpResponseRedirect(redirect_url)

# -------------------------- 4.5. 快速切换管理员状态（禁用/启用） --------------------------
@login_required
def user_toggle_admin_status(request, pk):
    """快速切换管理员的激活状态（无需进入编辑页）"""
    admin_user = get_object_or_404(User, pk=pk)
    # 验证是否是管理员
    if not admin_user.groups.filter(name='设备管理员').exists():
        messages.error(request, '该用户不是设备管理员！')
        return redirect('user_manage')
    
    admin_user.is_active = not admin_user.is_active  # 取反：正常→禁用，禁用→正常
    admin_user.save()
    
    messages.success(request, f'管理员【{admin_user.first_name or admin_user.username}】状态已更新为{"正常" if admin_user.is_active else "禁用"}')
    # 保留查询参数，避免跳转错误
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    redirect_url = reverse('user_manage')
    query_params = request.GET.copy()
    query_params['user_type'] = 'admin'  # 保持管理员筛选
    if query_params:
        redirect_url += '?' + query_params.urlencode()
    return HttpResponseRedirect(redirect_url)

# -------------------------- 5. 导出用户台账 --------------------------
@login_required
def user_export_ledger(request):
    """导出用户台账为Excel文件（.xlsx）"""
    from django.http import HttpResponse
    from django.db.models import Count
    
    user_type = request.GET.get('user_type', '')
    keyword = request.GET.get('keyword', '')
    
    # 根据筛选条件获取用户
    users = UserInfo.objects.annotate(booking_count=Count('booking'))
    if user_type and user_type in ['student', 'teacher', 'external']:
        users = users.filter(user_type=user_type)
    if keyword:
        users = users.filter(Q(name__icontains=keyword) | Q(user_code__icontains=keyword))
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "用户台账"
    
    # 根据用户类型设置不同的表头
    if user_type == 'teacher':
        headers = ['教师编号', '姓名', '性别', '职称', '专业方向', '所在学院', '联系电话', '借用次数', '借用资格', '创建时间']
        create_time_col = 10
    elif user_type == 'student':
        headers = ['学号', '姓名', '性别', '专业', '导师', '所在学院', '联系电话', '借用次数', '借用资格', '创建时间']
        create_time_col = 10
    elif user_type == 'external':
        headers = ['编号', '姓名', '性别', '所在单位名称', '联系电话', '借用次数', '借用资格', '创建时间']
        create_time_col = 8
    else:
        headers = ['用户编号', '姓名', '用户类型', '性别', '所在学院/单位', '联系电话', '借用次数', '借用资格', '创建时间']
        create_time_col = 9
    
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for user in users:
        # 转换datetime为naive datetime（移除时区信息）
        create_time = user.create_time.replace(tzinfo=None) if user.create_time else None
        
        if user_type == 'teacher':
            row = [
                user.user_code,
                user.name,
                user.gender,
                user.title or '-',
                user.research_field or '-',
                user.department,
                user.phone,
                user.booking_count,
                '正常' if user.is_active else '禁用',
                create_time,
            ]
        elif user_type == 'student':
            row = [
                user.user_code,
                user.name,
                user.gender,
                user.major or '-',
                ', '.join([a.name for a in user.advisors.all()]) if user.user_type == 'student' else '-',
                user.department,
                user.phone,
                user.booking_count,
                '正常' if user.is_active else '禁用',
                create_time,
            ]
        elif user_type == 'external':
            row = [
                user.user_code,
                user.name,
                user.gender,
                user.department,
                user.phone,
                user.booking_count,
                '正常' if user.is_active else '禁用',
                create_time,
            ]
        else:
            row = [
                user.user_code,
                user.name,
                user.get_user_type_display(),
                user.gender,
                user.department,
                user.phone,
                user.booking_count,
                '正常' if user.is_active else '禁用',
                create_time,
            ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if create_time:
            ws.cell(row=current_row, column=create_time_col).number_format = 'yyyy-mm-dd hh:mm:ss'
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="user_ledger.xlsx"'
    
    wb.save(response)
    return response