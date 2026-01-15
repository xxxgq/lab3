from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from django.db.models import Q, Count
import csv
from .models import DeviceLedger
from devices.models import Device, DEVICE_STATUS
from user.models import UserInfo
from booking.models import Booking
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def check_ledger_permission(view_func):
    """权限检查装饰器：只允许设备管理员和实验室负责人访问台账"""
    def wrapper(request, *args, **kwargs):
        # 确保用户已登录
        if not request.user.is_authenticated:
            messages.error(request, '请先登录！')
            return redirect('user_login')
        
        is_admin = request.user.groups.filter(name='设备管理员').exists()
        is_manager = request.user.groups.filter(name='实验室负责人').exists()
        if not is_admin and not is_manager and not request.user.is_superuser:
            messages.error(request, '您无权访问台账模块！')
            # 尝试重定向到管理员首页，如果不存在则重定向到登录页
            try:
                return redirect('admin_home')
            except:
                return redirect('user_login')
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@check_ledger_permission
def ledger_home(request):
    """台账选择页面"""
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    context = {
        'is_admin': is_admin,
        'is_manager': is_manager,
    }
    return render(request, 'ledger/ledger_home.html', context)

def get_user_role_context(request):
    """辅助函数：获取用户角色信息"""
    # 多角色支持：如果使用了角色特定的session，从那里获取用户
    from jnu_lab_system.multi_role_session import get_role_from_path, get_user_from_role_session
    
    role = get_role_from_path(request.path)
    current_user = request.user
    
    # 如果路径匹配某个角色，尝试从角色特定的session中获取用户
    if role:
        role_user = get_user_from_role_session(request, role)
        if role_user:
            current_user = role_user
    
    is_admin = current_user.groups.filter(name='设备管理员').exists() if current_user.is_authenticated else False
    is_manager = current_user.groups.filter(name='实验室负责人').exists() if current_user.is_authenticated else False
    return {
        'is_admin': is_admin,
        'is_manager': is_manager,
    }

@login_required
@check_ledger_permission
def device_ledger_list(request):
    """设备台账列表视图：显示所有设备信息"""
    from datetime import date, timedelta
    devices = Device.objects.all().order_by('device_code')

    # 筛选
    device_code = request.GET.get('device_code')
    if device_code:
        devices = devices.filter(device_code__icontains=device_code)

    model = request.GET.get('model')
    if model:
        devices = devices.filter(model__icontains=model)

    manufacturer = request.GET.get('manufacturer')
    if manufacturer:
        devices = devices.filter(manufacturer__icontains=manufacturer)

    status = request.GET.get('status')
    if status:
        devices = devices.filter(status=status)

    # 为每个设备计算时段可用状态（今天和未来6天，共7天）
    from datetime import date, timedelta, datetime
    today = date.today()
    now = datetime.now()
    time_slots = ['上午', '下午', '全天']
    device_time_status = {}
    
    # 获取所有相关预约（一次性查询，提高性能）
    future_date = today + timedelta(days=6)
    active_bookings = Booking.objects.filter(
        device__in=devices,
        booking_date__gte=today,
        booking_date__lte=future_date,
        status__in=['pending', 'admin_approved', 'manager_approved', 'payment_pending', 'teacher_pending']
    ).select_related('device', 'applicant')
    
    # 构建设备时段状态字典
    for device in devices:
        device_time_status[device.id] = []
        for day_offset in range(7):
            check_date = today + timedelta(days=day_offset)
            is_today = (check_date == today)
            day_status = {
                'date': check_date,
                'date_str': check_date.strftime('%m-%d'),
                'is_today': is_today,
                'slots': {}
            }
            
            for slot in time_slots:
                # 查找该设备在该日期该时段的预约
                booking = active_bookings.filter(
                    device=device,
                    booking_date=check_date,
                    time_slot=slot
                ).first()
                
                if booking:
                    # 判断当前时段是否正在进行中（仅今天）
                    is_current = False
                    if is_today:
                        hour = now.hour
                        if slot == '上午' and 8 <= hour < 12:
                            is_current = True
                        elif slot == '下午' and 12 <= hour < 18:
                            is_current = True
                        elif slot == '全天':
                            is_current = True
                    
                    day_status['slots'][slot] = {
                        'available': False,
                        'booked_by': booking.applicant.name,
                        'status': booking.get_status_display(),
                        'is_current': is_current,
                        'booking_code': booking.booking_code
                    }
                else:
                    # 判断当前时段是否可用（仅今天）
                    is_current = False
                    is_available_now = False
                    if is_today:
                        hour = now.hour
                        if slot == '上午' and 8 <= hour < 12:
                            is_current = True
                            is_available_now = True
                        elif slot == '下午' and 12 <= hour < 18:
                            is_current = True
                            is_available_now = True
                        elif slot == '全天':
                            is_current = True
                            is_available_now = True
                    
                    day_status['slots'][slot] = {
                        'available': True,
                        'booked_by': None,
                        'status': None,
                        'is_current': is_current,
                        'is_available_now': is_available_now
                    }
            
            device_time_status[device.id].append(day_status)

    # 分页
    paginator = Paginator(devices, 20)  # 每页20条记录
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'status_choices': DEVICE_STATUS,
        'total_count': devices.count(),
        'device_time_status': device_time_status,
        'today': today,
        'time_slots': time_slots,
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/device_info_ledger_list.html', context)

@login_required
@check_ledger_permission
def device_operation_history_list(request):
    """设备操作历史列表视图（保留原有功能）"""
    ledgers = DeviceLedger.objects.select_related('device', 'user', 'operator').order_by('-operation_date')

    # 筛选
    device_code = request.GET.get('device_code')
    if device_code:
        ledgers = ledgers.filter(device__device_code__icontains=device_code)

    operation_type = request.GET.get('operation_type')
    if operation_type:
        ledgers = ledgers.filter(operation_type=operation_type)

    date_from = request.GET.get('date_from')
    if date_from:
        ledgers = ledgers.filter(operation_date__date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        ledgers = ledgers.filter(operation_date__date__lte=date_to)

    operator_name = request.GET.get('operator')
    if operator_name:
        ledgers = ledgers.filter(operator__username__icontains=operator_name)

    # 分页
    paginator = Paginator(ledgers, 20)  # 每页20条记录
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'operation_types': DeviceLedger.OPERATION_TYPES,
        'total_count': ledgers.count(),
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/device_ledger_list.html', context)

@login_required
@check_ledger_permission
def teacher_ledger_list(request):
    """教师台账列表视图：显示申请过设备借用的教师信息"""
    # 筛选出申请过设备借用的教师（通过Booking关联），预加载设备信息
    teachers = UserInfo.objects.filter(
        user_type='teacher',
        booking__isnull=False
    ).distinct().prefetch_related('booking_set__device').annotate(
        booking_count=Count('booking')
    ).order_by('user_code')

    # 筛选
    user_code = request.GET.get('user_code')
    if user_code:
        teachers = teachers.filter(user_code__icontains=user_code)

    name = request.GET.get('name')
    if name:
        teachers = teachers.filter(name__icontains=name)

    department = request.GET.get('department')
    if department:
        teachers = teachers.filter(department__icontains=department)

    title = request.GET.get('title')
    if title:
        teachers = teachers.filter(title__icontains=title)

    # 分页
    paginator = Paginator(teachers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_count': teachers.count(),
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/teacher_ledger_list.html', context)

@login_required
@check_ledger_permission
def student_ledger_list(request):
    """学生台账列表视图：显示申请过设备借用的学生信息"""
    # 筛选出申请过设备借用的学生（通过Booking关联），预加载设备信息
    students = UserInfo.objects.filter(
        user_type='student',
        booking__isnull=False
    ).distinct().prefetch_related('booking_set__device').annotate(
        booking_count=Count('booking')
    ).order_by('user_code')

    # 筛选
    user_code = request.GET.get('user_code')
    if user_code:
        students = students.filter(user_code__icontains=user_code)

    name = request.GET.get('name')
    if name:
        students = students.filter(name__icontains=name)

    department = request.GET.get('department')
    if department:
        students = students.filter(department__icontains=department)

    major = request.GET.get('major')
    if major:
        students = students.filter(major__icontains=major)

    advisor = request.GET.get('advisor')
    if advisor:
        # 通过多对多关系查询指导教师
        students = students.filter(advisors__name__icontains=advisor).distinct()

    # 分页
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_count': students.count(),
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/student_ledger_list.html', context)

@login_required
@check_ledger_permission
def external_ledger_list(request):
    """校外人员台账列表视图：显示申请过设备借用的校外人员信息"""
    # 筛选出申请过设备借用的校外人员（通过Booking关联），预加载设备信息
    externals = UserInfo.objects.filter(
        user_type='external',
        booking__isnull=False
    ).distinct().prefetch_related('booking_set__device').annotate(
        booking_count=Count('booking')
    ).order_by('user_code')

    # 筛选
    user_code = request.GET.get('user_code')
    if user_code:
        externals = externals.filter(user_code__icontains=user_code)

    name = request.GET.get('name')
    if name:
        externals = externals.filter(name__icontains=name)

    department = request.GET.get('department')
    if department:
        externals = externals.filter(department__icontains=department)

    # 分页
    paginator = Paginator(externals, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_count': externals.count(),
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/external_ledger_list.html', context)

@login_required
@check_ledger_permission
def booking_ledger_list(request):
    """预约台账列表视图：显示所有预约申请信息"""
    bookings = Booking.objects.select_related('applicant', 'device').order_by('-create_time')

    # 筛选
    booking_code = request.GET.get('booking_code')
    if booking_code:
        bookings = bookings.filter(booking_code__icontains=booking_code)

    device_code = request.GET.get('device_code')
    if device_code:
        bookings = bookings.filter(device__device_code__icontains=device_code)

    applicant_name = request.GET.get('applicant_name')
    if applicant_name:
        bookings = bookings.filter(applicant__name__icontains=applicant_name)

    user_type = request.GET.get('user_type')
    if user_type:
        bookings = bookings.filter(applicant__user_type=user_type)

    status = request.GET.get('status')
    if status:
        bookings = bookings.filter(status=status)

    date_from = request.GET.get('date_from')
    if date_from:
        bookings = bookings.filter(booking_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        bookings = bookings.filter(booking_date__lte=date_to)

    # 分页
    paginator = Paginator(bookings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'status_choices': Booking.APPROVAL_STATUS,
        'user_type_choices': UserInfo.USER_TYPE_CHOICES,
        'total_count': bookings.count(),
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/booking_ledger_list.html', context)

@login_required
def device_ledger_detail(request, pk):
    """设备台账详情视图"""
    ledger = get_object_or_404(DeviceLedger, pk=pk)
    context = {
        'ledger': ledger,
    }
    context.update(get_user_role_context(request))
    return render(request, 'ledger/device_ledger_detail.html', context)

@login_required
@check_ledger_permission
def export_device_ledger_csv(request):
    """导出设备台账为Excel文件（.xlsx）"""
    
    devices = Device.objects.all().order_by('device_code')

    # 应用相同的筛选条件
    device_code = request.GET.get('device_code')
    if device_code:
        devices = devices.filter(device_code__icontains=device_code)

    model = request.GET.get('model')
    if model:
        devices = devices.filter(model__icontains=model)

    manufacturer = request.GET.get('manufacturer')
    if manufacturer:
        devices = devices.filter(manufacturer__icontains=manufacturer)

    status = request.GET.get('status')
    if status:
        devices = devices.filter(status=status)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "设备台账"

    # 设置表头
    headers = [
        '设备编号', '型号', '购入时间', '生产厂商', '实验用途', 
        '时段可用状态', '校内租用价格（元/2小时）', '校外租用价格（元/2小时）', '创建时间', '更新时间'
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期列的索引（从1开始，表头是第1行，数据从第2行开始）
    purchase_date_col = 3  # 购入时间
    created_at_col = 9     # 创建时间
    updated_at_col = 10    # 更新时间

    # 写入数据
    for device in devices:
        # 转换datetime为naive datetime（移除时区信息）
        purchase_date = device.purchase_date.date() if device.purchase_date else None
        created_at = device.created_at.replace(tzinfo=None) if device.created_at else None
        updated_at = device.updated_at.replace(tzinfo=None) if device.updated_at else None
        
        row = [
            device.device_code,
            device.model,
            purchase_date,
            device.manufacturer,
            device.purpose or '-',
            device.get_status_display(),
            device.price_internal,
            device.price_external,
            created_at,
            updated_at,
        ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if purchase_date:
            cell = ws.cell(row=current_row, column=purchase_date_col)
            cell.number_format = 'yyyy-mm-dd'
        if created_at:
            cell = ws.cell(row=current_row, column=created_at_col)
            cell.number_format = 'yyyy-mm-dd hh:mm:ss'
        if updated_at:
            cell = ws.cell(row=current_row, column=updated_at_col)
            cell.number_format = 'yyyy-mm-dd hh:mm:ss'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="device_info_ledger.xlsx"'
    
    wb.save(response)
    return response

@login_required
@check_ledger_permission
def export_teacher_ledger_csv(request):
    """导出教师台账为Excel文件（.xlsx）"""
    teachers = UserInfo.objects.filter(
        user_type='teacher',
        booking__isnull=False
    ).distinct().prefetch_related('booking_set__device').annotate(
        booking_count=Count('booking')
    ).order_by('user_code')

    # 应用相同的筛选条件
    user_code = request.GET.get('user_code')
    if user_code:
        teachers = teachers.filter(user_code__icontains=user_code)

    name = request.GET.get('name')
    if name:
        teachers = teachers.filter(name__icontains=name)

    department = request.GET.get('department')
    if department:
        teachers = teachers.filter(department__icontains=department)

    title = request.GET.get('title')
    if title:
        teachers = teachers.filter(title__icontains=title)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "教师台账"

    # 设置表头
    headers = ['教师编号', '姓名', '性别', '职称', '专业方向', '所在学院', '联系电话', '借用设备', '创建时间']
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期列的索引
    create_time_col = 9

    # 写入数据
    for teacher in teachers:
        device_codes = [booking.device.device_code for booking in teacher.booking_set.all()]
        device_str = '、'.join(device_codes) if device_codes else '-'
        # 转换datetime为naive datetime（移除时区信息）
        create_time = teacher.create_time.replace(tzinfo=None) if teacher.create_time else None
        
        row = [
            teacher.user_code,
            teacher.name,
            teacher.gender,
            teacher.title or '-',
            teacher.research_field or '-',
            teacher.department,
            teacher.phone,
            device_str,
            create_time,
        ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if create_time:
            ws.cell(row=current_row, column=create_time_col).number_format = 'yyyy-mm-dd hh:mm:ss'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="teacher_ledger.xlsx"'
    
    wb.save(response)
    return response

@login_required
@check_ledger_permission
def export_student_ledger_csv(request):
    """导出学生台账为Excel文件（.xlsx）"""
    students = UserInfo.objects.filter(
        user_type='student',
        booking__isnull=False
    ).distinct().prefetch_related('booking_set__device').annotate(
        booking_count=Count('booking')
    ).order_by('user_code')

    # 应用相同的筛选条件
    user_code = request.GET.get('user_code')
    if user_code:
        students = students.filter(user_code__icontains=user_code)

    name = request.GET.get('name')
    if name:
        students = students.filter(name__icontains=name)

    department = request.GET.get('department')
    if department:
        students = students.filter(department__icontains=department)

    major = request.GET.get('major')
    if major:
        students = students.filter(major__icontains=major)

    advisor = request.GET.get('advisor')
    if advisor:
        # 通过多对多关系查询指导教师
        students = students.filter(advisors__name__icontains=advisor).distinct()

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "学生台账"

    # 设置表头
    headers = ['学号', '姓名', '性别', '专业', '导师', '所在学院', '联系电话', '借用设备', '创建时间']
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期列的索引
    create_time_col = 9

    # 写入数据
    for student in students:
        device_codes = [booking.device.device_code for booking in student.booking_set.all()]
        device_str = '、'.join(device_codes) if device_codes else '-'
        # 转换datetime为naive datetime（移除时区信息）
        create_time = student.create_time.replace(tzinfo=None) if student.create_time else None
        
        row = [
            student.user_code,
            student.name,
            student.gender,
            student.major or '-',
            ', '.join([a.name for a in student.advisors.all()]) or '-',
            student.department,
            student.phone,
            device_str,
            create_time,
        ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if create_time:
            ws.cell(row=current_row, column=create_time_col).number_format = 'yyyy-mm-dd hh:mm:ss'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_ledger.xlsx"'
    
    wb.save(response)
    return response

@login_required
@check_ledger_permission
def export_external_ledger_csv(request):
    """导出校外人员台账为Excel文件（.xlsx）"""
    externals = UserInfo.objects.filter(
        user_type='external',
        booking__isnull=False
    ).distinct().prefetch_related('booking_set__device').annotate(
        booking_count=Count('booking')
    ).order_by('user_code')

    # 应用相同的筛选条件
    user_code = request.GET.get('user_code')
    if user_code:
        externals = externals.filter(user_code__icontains=user_code)

    name = request.GET.get('name')
    if name:
        externals = externals.filter(name__icontains=name)

    department = request.GET.get('department')
    if department:
        externals = externals.filter(department__icontains=department)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "校外人员台账"

    # 设置表头
    headers = ['编号', '姓名', '性别', '所在单位名称', '联系电话', '借用设备', '创建时间']
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期列的索引
    create_time_col = 7

    # 写入数据
    for external in externals:
        device_codes = [booking.device.device_code for booking in external.booking_set.all()]
        device_str = '、'.join(device_codes) if device_codes else '-'
        # 转换datetime为naive datetime（移除时区信息）
        create_time = external.create_time.replace(tzinfo=None) if external.create_time else None
        
        row = [
            external.user_code,
            external.name,
            external.gender,
            external.department,
            external.phone,
            device_str,
            create_time,
        ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if create_time:
            ws.cell(row=current_row, column=create_time_col).number_format = 'yyyy-mm-dd hh:mm:ss'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="external_ledger.xlsx"'
    
    wb.save(response)
    return response

@login_required
@check_ledger_permission
def export_booking_ledger_csv(request):
    """导出预约台账为Excel文件（.xlsx）"""
    bookings = Booking.objects.select_related('applicant', 'device').order_by('-create_time')

    # 应用相同的筛选条件
    booking_code = request.GET.get('booking_code')
    if booking_code:
        bookings = bookings.filter(booking_code__icontains=booking_code)

    device_code = request.GET.get('device_code')
    if device_code:
        bookings = bookings.filter(device__device_code__icontains=device_code)

    applicant_name = request.GET.get('applicant_name')
    if applicant_name:
        bookings = bookings.filter(applicant__name__icontains=applicant_name)

    user_type = request.GET.get('user_type')
    if user_type:
        bookings = bookings.filter(applicant__user_type=user_type)

    status = request.GET.get('status')
    if status:
        bookings = bookings.filter(status=status)

    date_from = request.GET.get('date_from')
    if date_from:
        bookings = bookings.filter(booking_date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        bookings = bookings.filter(booking_date__lte=date_to)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "预约台账"

    # 设置表头
    headers = [
        '预约编号', '申请人编号', '申请人姓名', '申请人类型', '设备编号', '设备型号',
        '预约日期', '预约时段', '借用用途', '指导教师编号', '审批状态', '创建时间', '更新时间'
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期列的索引
    booking_date_col = 7
    create_time_col = 12
    update_time_col = 13

    # 写入数据
    for booking in bookings:
        # 转换datetime为naive datetime（移除时区信息）
        booking_date = booking.booking_date if booking.booking_date else None
        create_time = booking.create_time.replace(tzinfo=None) if booking.create_time else None
        update_time = booking.update_time.replace(tzinfo=None) if booking.update_time else None
        
        row = [
            booking.booking_code,
            booking.applicant.user_code,
            booking.applicant.name,
            booking.applicant.get_user_type_display(),
            booking.device.device_code,
            booking.device.model,
            booking_date,
            booking.time_slot,
            booking.purpose or '-',
            booking.teacher.user_code if booking.teacher else '-',
            booking.get_status_display(),
            create_time,
            update_time,
        ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if booking_date:
            ws.cell(row=current_row, column=booking_date_col).number_format = 'yyyy-mm-dd'
        if create_time:
            ws.cell(row=current_row, column=create_time_col).number_format = 'yyyy-mm-dd hh:mm:ss'
        if update_time:
            ws.cell(row=current_row, column=update_time_col).number_format = 'yyyy-mm-dd hh:mm:ss'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="booking_ledger.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_ledger_csv(request):
    """导出设备操作历史为Excel文件（.xlsx）"""
    ledgers = DeviceLedger.objects.select_related('device', 'user', 'operator').order_by('-operation_date')

    # 应用相同的筛选条件
    device_code = request.GET.get('device_code')
    if device_code:
        ledgers = ledgers.filter(device__device_code__icontains=device_code)

    operation_type = request.GET.get('operation_type')
    if operation_type:
        ledgers = ledgers.filter(operation_type=operation_type)

    date_from = request.GET.get('date_from')
    if date_from:
        ledgers = ledgers.filter(operation_date__date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        ledgers = ledgers.filter(operation_date__date__lte=date_to)

    operator_name = request.GET.get('operator')
    if operator_name:
        ledgers = ledgers.filter(operator__username__icontains=operator_name)

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "设备操作历史"

    # 设置表头
    headers = [
        '设备编号', '设备名称', '借用人', '操作类型', '操作日期',
        '预期归还时间', '实际归还时间', '设备状态', '操作员', '备注'
    ]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 日期列的索引
    operation_date_col = 5
    expected_return_date_col = 6
    actual_return_date_col = 7

    # 写入数据
    for ledger in ledgers:
        # 处理设备编号：优先使用device.device_code，否则从description中提取
        if ledger.device:
            device_code = ledger.device.device_code
        elif ledger.operation_type == 'discard' and '删除设备：' in (ledger.description or ''):
            # 从删除描述中提取设备编号
            desc = ledger.description or ''
            if '删除设备：' in desc:
                device_code = desc.split('删除设备：')[1].split(' - ')[0]
            else:
                device_code = ledger.device_name
        else:
            device_code = ledger.device_name
        
        # 转换datetime为naive datetime（移除时区信息）
        operation_date = ledger.operation_date.replace(tzinfo=None) if ledger.operation_date else None
        expected_return_date = ledger.expected_return_date.replace(tzinfo=None) if ledger.expected_return_date else None
        actual_return_date = ledger.actual_return_date.replace(tzinfo=None) if ledger.actual_return_date else None
        
        row = [
            device_code,
            ledger.device_name,
            ledger.user.name if ledger.user else '',
            ledger.get_operation_type_display(),
            operation_date,
            expected_return_date,
            actual_return_date,
            ledger.get_status_after_operation_display(),
            ledger.operator.username if ledger.operator else '系统',
            ledger.description or ''
        ]
        ws.append(row)
        
        # 设置日期格式（短日期格式）
        current_row = ws.max_row
        if operation_date:
            ws.cell(row=current_row, column=operation_date_col).number_format = 'yyyy-mm-dd hh:mm:ss'
        if expected_return_date:
            ws.cell(row=current_row, column=expected_return_date_col).number_format = 'yyyy-mm-dd hh:mm:ss'
        if actual_return_date:
            ws.cell(row=current_row, column=actual_return_date_col).number_format = 'yyyy-mm-dd hh:mm:ss'

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="device_operation_history.xlsx"'
    
    wb.save(response)
    return response