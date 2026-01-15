"""Excel批量导入学生列表功能"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.contrib.auth.models import User
from user.models import UserInfo
from openpyxl import load_workbook
from django.db import transaction

@login_required
@csrf_protect
def import_students_excel(request):
    """教师批量导入学生（Excel）"""
    # 获取当前教师信息
    try:
        teacher_info = UserInfo.objects.get(auth_user=request.user)
        # 严格检查：必须是教师类型
        if teacher_info.user_type != 'teacher':
            messages.error(request, '只有教师可以批量导入学生！')
            # 根据用户类型重定向到正确的首页，避免循环跳转
            if teacher_info.user_type == 'student':
                return redirect('user_home')
            elif teacher_info.user_type == 'external':
                return redirect('user_home')
            else:
                return redirect('user_home')
    except UserInfo.DoesNotExist:
        messages.error(request, '未找到你的个人信息，请联系管理员！')
        return redirect('user_home')
    
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, '请选择Excel文件！')
            return redirect('user_profile')
        
        excel_file = request.FILES['excel_file']
        
        # 验证文件类型
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, '请上传Excel文件（.xlsx或.xls格式）！')
            return redirect('user_profile')
        
        try:
            # 读取Excel文件
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # 验证表头（期望的列：学号、姓名、性别、专业、所在学院、联系电话）
            headers = []
            if ws.max_row > 0:
                headers = [cell.value for cell in ws[1]]
            
            # 查找列索引
            col_indices = {}
            expected_cols = ['学号', '姓名', '性别', '专业', '所在学院', '联系电话']
            for col in expected_cols:
                try:
                    col_indices[col] = headers.index(col)
                except ValueError:
                    messages.error(request, f'Excel文件缺少必需的列：{col}')
                    return redirect('user_profile')
            
            # 读取数据并创建学生
            success_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    # 跳过空行
                    if not any(cell.value for cell in row):
                        continue
                    
                    try:
                        # 获取数据
                        user_code = str(row[col_indices['学号']].value).strip() if row[col_indices['学号']].value else None
                        name = str(row[col_indices['姓名']].value).strip() if row[col_indices['姓名']].value else None
                        gender = str(row[col_indices['性别']].value).strip() if row[col_indices['性别']].value else '男'
                        major = str(row[col_indices['专业']].value).strip() if row[col_indices['专业']].value else ''
                        department = str(row[col_indices['所在学院']].value).strip() if row[col_indices['所在学院']].value else ''
                        phone = str(row[col_indices['联系电话']].value).strip() if row[col_indices['联系电话']].value else ''
                        
                        # 验证必填字段
                        if not user_code or not name:
                            errors.append(f'第{row_idx}行：学号和姓名不能为空')
                            error_count += 1
                            continue
                        
                        # 检查学号是否已存在
                        if UserInfo.objects.filter(user_code=user_code).exists():
                            # 如果已存在，添加到指导教师的多对多关系中
                            existing_student = UserInfo.objects.get(user_code=user_code)
                            if existing_student.user_type == 'student':
                                if teacher_info not in existing_student.advisors.all():
                                    existing_student.advisors.add(teacher_info)
                                success_count += 1
                            else:
                                errors.append(f'第{row_idx}行：学号 {user_code} 对应的用户不是学生')
                                error_count += 1
                            continue
                        
                        # 创建Django User
                        password = user_code  # 默认密码为学号
                        user = User.objects.create_user(
                            username=user_code,
                            password=password,
                            first_name=name,
                            is_active=True
                        )
                        
                        # 创建UserInfo
                        student = UserInfo.objects.create(
                            user_code=user_code,
                            name=name,
                            gender=gender,
                            user_type='student',
                            department=department,
                            phone=phone,
                            major=major,
                            auth_user=user,
                            is_active=True
                        )
                        
                        # 将学生添加到教师的多对多关系中
                        student.advisors.add(teacher_info)
                        
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f'第{row_idx}行：{str(e)}')
                        error_count += 1
                        continue
            
            # 显示结果
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 名学生！')
            if error_count > 0:
                error_msg = f'有 {error_count} 条记录导入失败：' + '; '.join(errors[:10])  # 最多显示10个错误
                if len(errors) > 10:
                    error_msg += f'...（还有 {len(errors) - 10} 个错误）'
                messages.warning(request, error_msg)
            
        except Exception as e:
            messages.error(request, f'导入失败：{str(e)}')
        
        return redirect('user_profile')
    
    # GET请求：显示导入页面
    return render(request, 'user/import_students_excel.html', {
        'teacher_info': teacher_info
    })
