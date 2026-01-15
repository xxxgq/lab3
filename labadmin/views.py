from django.shortcuts import render, redirect, get_object_or_404

from django.db.models import Q
from user.models import UserInfo
from user.forms import UserInfoForm
from django.contrib.auth.hashers import make_password  # 密码加密
from django.contrib.auth import update_session_auth_hash  # 保持登录状态

# 以下是创建角色组和初始用户的代码
from django.contrib.auth.models import User, Group, Permission

from django.contrib.auth.decorators import login_required
from django.contrib import messages

from booking.models import Booking, ApprovalRecord
from user.models import UserInfo
from devices.models import Device
from ledger.models import DeviceLedger
from .models import Report
from django.utils import timezone
from datetime import timedelta, datetime, date
from django.db.models import Count, Sum, Q, Avg
from django.http import JsonResponse, HttpResponse
import json
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


@login_required
def admin_home(request):
    """管理员首页 - 仅允许管理员访问"""
    # 验证用户身份：确保request.user是正确的用户
    if not request.user.is_authenticated:
        from django.contrib.auth import logout
        logout(request)
        messages.error(request, '登录已过期，请重新登录！')
        return redirect('user_login')
    
    # 验证用户是否是管理员（严格检查，不允许负责人访问）
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 关键修复：如果是负责人但不是管理员，必须重定向到负责人首页
    if is_manager and not is_admin and not request.user.is_superuser:
        messages.info(request, '您已切换到负责人界面')
        return redirect('manager_home')
    
    # 如果不是管理员，重定向到对应首页或登录页
    if not is_admin and not request.user.is_superuser:
        # 如果是普通用户，重定向到普通用户首页
        try:
            from user.models import UserInfo
            user_info = UserInfo.objects.get(auth_user=request.user)
            return redirect('user_home')
        except:
            pass
        # 否则重定向到登录页
        messages.error(request, '您不是设备管理员，无权访问此页面！')
        return redirect('user_login')
    
    context = {
        'is_admin': True,  # 明确设置为True，因为已经通过验证
        'is_manager': False,  # 管理员首页不应该显示负责人相关功能
    }
    return render(request, 'admin/home.html', context)

@login_required
def device_list(request):
    """
    用户端设备查询视图
    对应路径：/user/device/list/
    """
    # 1. 处理搜索逻辑
    keyword = request.GET.get('keyword', '')
    # 基础查询：获取所有设备（按编号排序）
    devices = Device.objects.all().order_by('device_code')
    
    # 如果有搜索关键词，过滤结果
    if keyword:
        devices = devices.filter(
            Q(device_code__icontains=keyword) |  # 按设备编号搜索
            Q(model__icontains=keyword) |        # 按型号搜索
            Q(manufacturer__icontains=keyword) | # 按厂商搜索
            Q(purpose__icontains=keyword)        # 按实验用途搜索
        )

    # 2. 准备上下文数据
    context = {
        'devices': devices,
        'keyword': keyword,  # 回显搜索关键词
    }
    return render(request, 'user/device_list.html', context)

# 注意：booking_apply 和 my_booking 视图函数已移至 booking/views.py
# 这里不再需要这些函数，避免与 booking/views.py 中的函数冲突

def generate_report_data(report_type, start_date, end_date):
    """生成报表数据"""
    # 获取时间范围内的预约数据
    bookings = Booking.objects.filter(
        booking_date__gte=start_date,
        booking_date__lte=end_date
    )
    
    # 获取已审批通过的预约
    approved_bookings = bookings.filter(status='manager_approved')
    
    # 基础统计
    total_bookings = bookings.count()
    approved_count = approved_bookings.count()
    rejected_count = bookings.filter(Q(status='admin_rejected') | Q(status='manager_rejected')).count()
    pending_count = bookings.filter(status='pending').count()
    
    # 按设备统计
    device_stats = approved_bookings.values('device__device_code', 'device__model').annotate(
        booking_count=Count('id'),
        revenue=Sum('device__price_external')
    ).order_by('-booking_count')
    
    # 按用户类型统计
    user_type_stats = approved_bookings.values('applicant__user_type').annotate(
        booking_count=Count('id'),
        user_count=Count('applicant', distinct=True)
    )
    
    # 按日期统计（用于图表）
    date_stats = approved_bookings.values('booking_date').annotate(
        booking_count=Count('id')
    ).order_by('booking_date')
    
    # 计算总收入（仅校外人员）
    total_revenue = approved_bookings.filter(
        applicant__user_type='external'
    ).aggregate(
        total=Sum('device__price_external')
    )['total'] or Decimal('0')
    
    # 设备使用率统计
    device_usage = []
    for device in Device.objects.all():
        device_bookings = approved_bookings.filter(device=device)
        booking_count = device_bookings.count()
        # 假设每个预约使用2小时
        usage_hours = booking_count * 2
        # 计算使用率（假设每天可用8小时）
        days = (end_date - start_date).days + 1
        total_hours = days * 8
        usage_rate = (usage_hours / total_hours * 100) if total_hours > 0 else 0
        
        device_usage.append({
            'device_code': device.device_code,
            'device_model': device.model,
            'booking_count': booking_count,
            'usage_hours': usage_hours,
            'usage_rate': round(usage_rate, 2),
            'revenue': float(device_bookings.filter(applicant__user_type='external').aggregate(
                total=Sum('device__price_external')
            )['total'] or Decimal('0'))
        })
    
    # 构建报表数据
    report_data = {
        'summary': {
            'total_bookings': total_bookings,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'pending_count': pending_count,
            'total_devices': Device.objects.count(),
            'total_users': UserInfo.objects.filter(booking__in=approved_bookings).distinct().count(),
            'total_revenue': float(total_revenue),
        },
        'device_stats': list(device_stats),
        'user_type_stats': list(user_type_stats),
        'date_stats': list(date_stats),
        'device_usage': device_usage,
    }
    
    return report_data

@login_required
def report_stat(request):
    """报表统计页面 - 仅允许管理员访问"""
    # 验证用户是否是管理员（严格检查，不允许负责人访问）
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 关键修复：如果是负责人但不是管理员，重定向到负责人报表页面
    if is_manager and not is_admin and not request.user.is_superuser:
        messages.info(request, '负责人请使用负责人报表页面')
        return redirect('manager_report_stat')
    
    # 如果不是管理员，重定向到对应首页或登录页
    if not is_admin and not request.user.is_superuser:
        # 如果是普通用户，重定向到普通用户首页
        try:
            from user.models import UserInfo
            user_info = UserInfo.objects.get(auth_user=request.user)
            return redirect('user_home')
        except:
            pass
        messages.error(request, '您不是设备管理员，无权访问此页面！')
        return redirect('user_login')
    
    # 获取已生成的报表列表
    reports = Report.objects.all().order_by('-generated_at')[:20]
    
    # 获取筛选条件
    report_type_filter = request.GET.get('report_type', '')
    date_filter = request.GET.get('date', '')
    
    if report_type_filter:
        reports = reports.filter(report_type=report_type_filter)
    
    # 处理报表生成请求
    if request.method == 'POST' and 'generate' in request.POST:
        report_type = request.POST.get('report_type')
        date_input = request.POST.get('date_input')
        start_date_input = request.POST.get('start_date', '').strip()
        end_date_input = request.POST.get('end_date', '').strip()
        
        # 自定义时间段报表需要起始日期和结束日期
        if report_type == 'custom':
            if not start_date_input or not end_date_input:
                messages.error(request, '自定义时间段报表需要填写起始日期和结束日期！')
                return redirect('report_stat')
        elif not date_input:
            messages.error(request, '请选择报表类型和日期！')
            return redirect('report_stat')
        
        try:
            # 解析日期
            if report_type == 'week':
                # 周报表：输入日期所在周的周一和周日
                input_date = datetime.strptime(date_input, '%Y-%m-%d').date()
                start_date = input_date - timedelta(days=input_date.weekday())
                end_date = start_date + timedelta(days=6)
                report_name = f"{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')} 周报表"
            elif report_type == 'month':
                # 月报表：输入日期所在月的第一天和最后一天
                # 处理 YYYY-MM 格式
                if len(date_input) == 7 and date_input.count('-') == 1:
                    year, month = map(int, date_input.split('-'))
                else:
                    # 尝试解析为日期
                    input_date = datetime.strptime(date_input, '%Y-%m-%d').date()
                    year, month = input_date.year, input_date.month
                start_date = date(year, month, 1)
                if month == 12:
                    end_date = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(year, month + 1, 1) - timedelta(days=1)
                report_name = f"{year}年{month:02d}月报表"
            elif report_type == 'year':
                # 年报表：输入日期所在年的1月1日和12月31日
                year = int(date_input)
                start_date = date(year, 1, 1)
                end_date = date(year, 12, 31)
                report_name = f"{year}年报表"
            elif report_type == 'custom':
                # 自定义时间段报表：使用用户指定的起始日期和结束日期
                start_date = datetime.strptime(start_date_input, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_input, '%Y-%m-%d').date()
                
                # 验证日期范围
                if start_date > end_date:
                    messages.error(request, '起始日期不能晚于结束日期！')
                    return redirect('report_stat')
                
                report_name = f"{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')} 自定义报表"
            else:
                messages.error(request, '无效的报表类型！')
                return redirect('report_stat')
            
            # 检查是否已存在相同报表（只检查相同类型的，自定义报表可以重复生成）
            if report_type != 'custom':
                existing_report = Report.objects.filter(
                    report_type=report_type,
                    start_date=start_date,
                    end_date=end_date
                ).first()
                
                if existing_report:
                    if existing_report.generated_by:
                        messages.info(request, f'该时间段报表已存在（手动生成），已为您加载：{existing_report.report_name}')
                    else:
                        messages.info(request, f'该时间段报表已存在（系统自动生成），已为您加载：{existing_report.report_name}')
                    return redirect(f'report_stat?view={existing_report.id}')
            
            # 生成报表数据
            report_data = generate_report_data(report_type, start_date, end_date)
            
            # 创建报表记录
            report = Report.objects.create(
                report_type=report_type,
                report_name=report_name,
                start_date=start_date,
                end_date=end_date,
                report_data=report_data,
                total_bookings=report_data['summary']['total_bookings'],
                total_devices=report_data['summary']['total_devices'],
                total_users=report_data['summary']['total_users'],
                total_revenue=Decimal(str(report_data['summary']['total_revenue'])),
                generated_by=request.user
            )
            
            messages.success(request, f'报表生成成功：{report_name}')
            return redirect(f'report_stat?view={report.id}')
            
        except ValueError as e:
            messages.error(request, f'日期格式错误：请检查日期格式是否正确！')
            return redirect('report_stat')
        except Exception as e:
            messages.error(request, f'生成报表失败：{str(e)}')
            return redirect('report_stat')
    
    # 查看报表详情
    report_id = request.GET.get('view')
    current_report = None
    if report_id:
        try:
            current_report = Report.objects.get(id=report_id)
        except Report.DoesNotExist:
            messages.error(request, '报表不存在！')
    
    # 获取用户角色信息
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 处理报表删除请求
    if request.method == 'POST' and 'delete_report' in request.POST:
        delete_report_id = request.POST.get('delete_report')
        try:
            report = Report.objects.get(id=delete_report_id)
            report_name = report.report_name
            report.delete()
            messages.success(request, f'报表【{report_name}】已成功删除！')
            # 如果删除的是当前查看的报表，清除view参数
            if str(report_id) == str(delete_report_id):
                return redirect('report_stat')
        except Report.DoesNotExist:
            messages.error(request, '报表不存在！')
        except Exception as e:
            messages.error(request, f'删除报表失败：{str(e)}')
        # 保留查询参数
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        redirect_url = reverse('report_stat')
        query_params = request.GET.copy()
        if 'view' in query_params and str(query_params['view']) == str(delete_report_id):
            del query_params['view']
        if query_params:
            redirect_url += '?' + query_params.urlencode()
        return HttpResponseRedirect(redirect_url)
    
    context = {
        'reports': reports,
        'current_report': current_report,
        'report_type_filter': report_type_filter,
        'date_filter': date_filter,
        'is_admin': is_admin,
        'is_manager': is_manager,
    }
    return render(request, 'admin/report_stat.html', context)

@login_required
def delete_report(request, report_id):
    """删除报表（GET方式，带确认）"""
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    
    try:
        report = Report.objects.get(id=report_id)
        report_name = report.report_name
        report.delete()
        messages.success(request, f'报表【{report_name}】已成功删除！')
    except Report.DoesNotExist:
        messages.error(request, '报表不存在！')
    except Exception as e:
        messages.error(request, f'删除报表失败：{str(e)}')
    
    # 保留查询参数，避免跳转错误
    redirect_url = reverse('report_stat')
    query_params = request.GET.copy()
    if 'view' in query_params:
        del query_params['view']  # 删除view参数，因为报表已删除
    if query_params:
        redirect_url += '?' + query_params.urlencode()
    return HttpResponseRedirect(redirect_url)

@login_required
def export_report_csv(request, report_id):
    """导出报表为Excel文件（.xlsx）"""
    from django.http import HttpResponse
    from .models import Report
    import re
    
    report = get_object_or_404(Report, id=report_id)
    data = report.get_report_data()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "报表"
    
    # 写入报表基本信息
    # 转换datetime为naive datetime（移除时区信息）
    generated_at = report.generated_at.replace(tzinfo=None) if report.generated_at else None
    
    ws.append(['报表名称', report.report_name])
    ws.append(['报表类型', report.get_report_type_display()])
    ws.append(['统计时间', f'{report.start_date.strftime("%Y-%m-%d")} 至 {report.end_date.strftime("%Y-%m-%d")}'])
    ws.append(['生成时间', generated_at])
    ws.append(['生成人', report.generated_by.username if report.generated_by else '系统自动'])
    ws.append([])  # 空行
    
    # 设置生成时间为日期格式
    if generated_at:
        ws.cell(row=4, column=2).number_format = 'yyyy-mm-dd hh:mm:ss'
    
    # 写入汇总统计
    ws.append(['汇总统计'])
    ws.append(['总预约次数', data['summary']['total_bookings']])
    ws.append(['已审批通过', data['summary']['approved_count']])
    ws.append(['总收入（元）', data['summary']['total_revenue']])
    ws.append(['设备总数', data['summary']['total_devices']])
    ws.append(['用户总数', data['summary']['total_users']])
    ws.append([])  # 空行
    
    # 写入设备使用统计
    ws.append(['设备使用统计'])
    headers = ['设备编号', '设备型号', '预约次数', '使用时长（小时）', '使用率（%）', '校外收费（元）']
    ws.append(headers)
    
    # 设置表头样式
    header_row = ws.max_row
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for device in data.get('device_usage', []):
        ws.append([
            device['device_code'],
            device['device_model'],
            device['booking_count'],
            device['usage_hours'],
            f"{device['usage_rate']}%",
            device['revenue']
        ])
    ws.append([])  # 空行
    
    # 写入用户类型统计
    ws.append(['用户类型统计'])
    headers = ['用户类型', '预约次数', '用户数量']
    ws.append(headers)
    
    # 设置表头样式
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for stat in data.get('user_type_stats', []):
        user_type = stat.get('applicant__user_type', '')
        user_type_display = {
            'student': '校内学生',
            'teacher': '校内教师',
            'external': '校外人员'
        }.get(user_type, user_type)
        ws.append([
            user_type_display,
            stat['booking_count'],
            stat['user_count']
        ])
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 创建响应
    # 清理文件名中的特殊字符
    safe_filename = re.sub(r'[<>:"/\\|?*]', '_', report.report_name)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{report.id}_{safe_filename}.xlsx"'
    
    wb.save(response)
    return response

# 1. 管理员审批页面
@login_required
def booking_approve(request):
    """设备预约审批（管理员） - 仅允许管理员访问"""
    # 校验是否是管理员（严格检查，不允许负责人访问管理员审批页面）
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 关键修复：如果是负责人但不是管理员，重定向到负责人审批页面
    if is_manager and not is_admin and not request.user.is_superuser:
        messages.info(request, '负责人请使用负责人审批页面')
        return redirect('manager_booking_approve')
    
    # 如果不是管理员，重定向到对应首页或登录页
    if not is_admin and not request.user.is_superuser:
        # 如果是普通用户，重定向到普通用户首页
        try:
            from user.models import UserInfo
            user_info = UserInfo.objects.get(auth_user=request.user)
            return redirect('user_home')
        except:
            pass
        messages.error(request, '你无审批权限！')
        return redirect('user_login')
    
    # 获取筛选条件
    user_type_filter = request.GET.get('user_type', 'all')
    
    # 核心修改：管理员只看「待审批（pending）」的申请，不看「已批准待负责人审批（admin_approved）」的
    if is_admin:
        # 管理员仅审批：待审批的申请（移除 admin_approved）
        bookings = Booking.objects.filter(
            status='pending'  # 只保留待管理员审批的
        ).order_by('-create_time')
    else:
        # 负责人仍只审批：管理员已批准的校外人员申请（admin_approved + external）
        bookings = Booking.objects.filter(
            status='admin_approved',
            applicant__user_type='external'
        ).order_by('-create_time')
    
    # 筛选用户类型
    if user_type_filter != 'all':
        if user_type_filter == 'student':
            bookings = bookings.filter(applicant__user_type='student')
        elif user_type_filter == 'teacher':
            bookings = bookings.filter(applicant__user_type='teacher')
        elif user_type_filter == 'external':
            bookings = bookings.filter(applicant__user_type='external')
    
    # 处理审批操作（原有逻辑不变）
    if request.method == 'POST':
        # 单个审批
        if 'approve' in request.POST:
            booking_id = request.POST.get('approve')
            handle_approval(request, booking_id, 'approve')
        elif 'reject' in request.POST:
            booking_id = request.POST.get('reject')
            handle_approval(request, booking_id, 'reject')
        
        # 批量审批（简化版）
        elif 'batch_approve' in request.POST or 'batch_reject' in request.POST:
            booking_ids = request.POST.getlist('booking_ids')
            action = 'approve' if 'batch_approve' in request.POST else 'reject'
            for booking_id in booking_ids:
                handle_approval(request, booking_id, action)
        
        # 保留查询参数（筛选条件），避免跳转错误
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        redirect_url = reverse('booking_approve')
        query_params = request.GET.copy()
        if query_params:
            redirect_url += '?' + query_params.urlencode()
        return HttpResponseRedirect(redirect_url)
    
    context = {
        'bookings': bookings,
        'user_type_filter': user_type_filter,
        'is_admin': is_admin,
        'is_manager': is_manager
    }
    return render(request, 'admin/booking_approve.html', context)

def handle_approval(request, booking_id, action):
    """处理审批逻辑（核心）"""
    booking = get_object_or_404(Booking, id=booking_id)
    is_admin = request.user.groups.filter(name='设备管理员').exists()
    is_manager = request.user.groups.filter(name='实验室负责人').exists()
    
    # 1. 管理员审批逻辑
    if is_admin:
        if action == 'approve':
            # 学生/教师：直接审批通过
            if booking.applicant.user_type in ['student', 'teacher']:
                booking.status = 'manager_approved'
                # 审批通过时创建借出台账记录
                create_borrow_ledger(booking, request.user)
            # 校外人员：需负责人审批
            else:
                booking.status = 'admin_approved'
            approval_level = 'admin'
        else:
            booking.status = 'admin_rejected'
            approval_level = 'admin'
    
    # 2. 负责人审批逻辑（仅校外人员）
    elif is_manager:
        if action == 'approve':
            # 校外人员：负责人批准后，需要缴费
            booking.status = 'payment_pending'
            # 发送缴费请求到财务处
            from booking.finance_integration import send_payment_request_to_finance
            payment_result = send_payment_request_to_finance(booking)
            # 注意：实际缴费确认需要通过财务处回调接口完成
        else:
            booking.status = 'manager_rejected'
        approval_level = 'manager'
    
    # 保存预约状态
    booking.save()
    
    # 记录审批日志
    ApprovalRecord.objects.create(
        booking=booking,
        approver=request.user,
        approval_level=approval_level,
        action=action,
        comment=request.POST.get(f'comment_{booking.booking_code}', '')  # 可扩展审批备注
    )
    
    # 提示信息
    action_text = '批准' if action == 'approve' else '拒绝'
    messages.success(request, f'已{action_text}预约申请：{booking.booking_code}')

def create_borrow_ledger(booking, operator):
    """审批通过时创建借出台账记录"""
    try:
        from datetime import datetime, time as dt_time
        
        # 计算预期归还时间（基于预约日期和时段）
        # 解析时段，获取结束时间
        time_slot = booking.time_slot
        end_time_str = time_slot.split('-')[1] if '-' in time_slot else '20:00'
        end_hour, end_minute = map(int, end_time_str.split(':'))
        
        # 创建预期归还时间（预约日期的结束时间）
        expected_return_date = datetime.combine(
            booking.booking_date,
            dt_time(end_hour, end_minute)
        )
        
        # 判断设备状态：只有在预约日期是今天且当前时段正在进行时才设置为"不可用"
        # 否则保持设备当前状态，由定时任务来管理
        now = timezone.now()
        today = now.date()
        current_time = now.time()
        
        # 解析时段开始时间
        start_time_str = time_slot.split('-')[0] if '-' in time_slot else '08:00'
        start_hour, start_minute = map(int, start_time_str.split(':'))
        slot_start_time = dt_time(start_hour, start_minute)
        slot_end_time = dt_time(end_hour, end_minute)
        
        # 创建借出台账记录
        # 注意：设备状态不再设置为"不可用"，时段可用性由预约情况决定
        DeviceLedger.objects.create(
            device=booking.device,
            device_name=booking.device.model,
            user=booking.applicant,
            operation_type='borrow',
            operation_date=timezone.now(),
            expected_return_date=expected_return_date,
            status_after_operation=booking.device.status,  # 保持设备当前状态（正常/维修中/已报废）
            description=f'预约编号：{booking.booking_code}，用途：{booking.purpose or "无"}',
            operator=operator
        )
        
        # 设备状态不再因为预约而改变，时段可用性通过预约情况来判断
        print(f'已为预约 {booking.booking_code} 创建借出台账记录（预约日期：{booking.booking_date}，时段：{booking.time_slot}）')
        
    except Exception as e:
        print(f'创建台账记录失败：{str(e)}')
